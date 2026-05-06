import hmac, os, re
import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- PASSWORD --
def check_password():
    def _entered():
        try:
            correct = st.secrets.get("APP_PASSWORD", "")
        except Exception:
            correct = os.environ.get("APP_PASSWORD", "")
        if correct and hmac.compare_digest(st.session_state["pw"], correct):
            st.session_state["auth"] = True
            del st.session_state["pw"]
        else:
            st.session_state["auth"] = False
    if st.session_state.get("auth"):
        return True
    st.markdown("## SDR Generator")
    st.text_input("Password", type="password", on_change=_entered, key="pw")
    if st.session_state.get("auth") == False:
        st.error("Incorrect password.")
    return False

_pw = False
try:
    _pw = bool(st.secrets.get("APP_PASSWORD", ""))
except Exception:
    _pw = bool(os.environ.get("APP_PASSWORD", ""))
if _pw and not check_password():
    st.stop()

# -- HELPERS --
def slug(t):
    return re.sub(r"[^a-z0-9]+", "_", t.lower().strip()).strip("_")

def _safe_loc(loc, el_type="link"):
    VALID_EXIT_LOCS     = {"header", "footer", "inpage", "sidebar", "popup"}
    VALID_DOWNLOAD_LOCS = {"header", "footer", "inpage", "sidebar", "hero"}
    if el_type == "exit":
        if loc == "modal":
            return "popup"
        return loc if loc in VALID_EXIT_LOCS else "inpage"
    if el_type == "download":
        return loc if loc in VALID_DOWNLOAD_LOCS else "inpage"
    return loc

def detect_location(tag):
    for parent in tag.parents:
        if parent.name is None:
            continue
        tag_name = parent.name.lower()
        classes  = " ".join(parent.get("class", [])).lower()
        pid      = (parent.get("id") or "").lower()
        combined = f"{tag_name} {classes} {pid}"
        if tag_name == "header" or any(x in combined for x in
                ["header","site-header","navbar","nav-top","top-nav","masthead","utility-nav","utility-bar"]):
            return "header"
        if tag_name == "footer" or any(x in combined for x in
                ["footer","site-footer","footer-nav","foot-"]):
            return "footer"
        if tag_name == "nav" or any(x in combined for x in
                ["nav","navigation","main-nav","primary-nav","menu"]):
            return "header"
        if any(x in combined for x in ["hero","banner","jumbotron","splash","carousel"]):
            return "hero"
        if any(x in combined for x in ["modal","dialog","popup","overlay","lightbox"]):
            return "popup"
        if any(x in combined for x in ["sidebar","side-bar","aside","widget"]):
            return "sidebar"
    return "inpage"

def _resolve_video_source(el):
    for attr in ["src", "data-src", "data-url", "data-video-url",
                 "data-video-src", "data-embed-url"]:
        v = (el.get(attr) or "").lower()
        if "youtube" in v or "youtu.be" in v:
            return "youtube", el.get(attr, "")
        if "vimeo"      in v: return "vimeo",      el.get(attr, "")
        if "wistia"     in v: return "wistia",     el.get(attr, "")
        if "brightcove" in v: return "brightcove", el.get(attr, "")
    return "html5", el.get("src", "")

def _infer_cta_type(tag):
    if tag is None:
        return "link"
    tag_name = tag.name.lower() if tag.name else ""
    classes  = " ".join(tag.get("class", [])).lower()
    pid      = (tag.get("id") or "").lower()
    combined = f"{tag_name} {classes} {pid}"
    role     = tag.get("role", "").lower()
    if any(x in combined for x in ["logo","brand","site-logo","navbar-brand","header-logo"]):
        return "logo"
    img = tag.find("img")
    if img:
        img_alt = (img.get("alt","") or "").lower()
        img_src = (img.get("src","") or "").lower()
        if any(x in img_alt + img_src for x in ["logo","brand","home"]):
            return "logo"
        return "image_link"
    if role == "button":
        return "button"
    if any(x in combined for x in ["btn","button","cta","call-to-action","action-btn"]):
        return "button"
    if any(x in combined for x in ["nav","menu","navigation","breadcrumb"]):
        return "nav_link"
    for parent in tag.parents:
        if parent.name in ("nav", "header"):
            return "nav_link"
        if parent.name and any(x in " ".join(parent.get("class",[])).lower()
                               for x in ["nav","menu","navigation"]):
            return "nav_link"
        if parent.name in ("footer",):
            return "footer_link"
        break
    if not tag.get_text(strip=True) and tag.find("img"):
        return "image_link"
    return "link"

# -- QUICK SCAN --
def quick_scan(url, page_name="Scanned"):
    from urllib.parse import urlparse, urljoin
    parsed = urlparse(url)
    auth = None
    if parsed.username and parsed.password:
        auth = (parsed.username, parsed.password)
        port_part = f":{parsed.port}" if parsed.port else ""
        clean_url = parsed._replace(netloc=parsed.hostname + port_part).geturl()
    else:
        clean_url = url
    parsed = urlparse(clean_url)

    html = None
    try:
        from curl_cffi import requests as cr
        r = cr.get(clean_url, impersonate="chrome110", timeout=20, auth=auth)
        if r.status_code == 200:
            html = r.text
    except Exception:
        pass
    if not html:
        try:
            import requests
            r = requests.get(clean_url, timeout=15, auth=auth,
                             headers={"User-Agent": "Mozilla/5.0 Chrome/120"})
            if r.status_code == 200:
                html = r.text
        except Exception:
            pass
    if not html:
        return []

    from bs4 import BeautifulSoup
    soup    = BeautifulSoup(html, "lxml")
    rows    = []
    dl_exts = {".pdf", ".docx", ".xlsx", ".zip", ".ppt", ".pptx"}

    seen_events = set()
    def _add(row):
        key = (row["Event Name"], row["Label"], row["Page URL"])
        if key in seen_events:
            return
        seen_events.add(key)
        rows.append(row)

    _page_slug = slug(clean_url.rstrip("/").split("/")[-1] or "home")[:20]

    # 1. SCROLL EVENTS
    for pct in [25, 50, 75, 100]:
        _add({
            "Page": page_name, "Category": "scroll",
            "Event Name": f"scroll_depth_{pct}", "Action": "scroll",
            "Label": f"{pct}%", "CTA Location": "inpage",
            "CTA Text": f"Scroll {pct}%", "Business Intent": "engagement",
            "Page URL": clean_url, "Event ID": "",
            "Form Info": "", "Video Source": "", "File Type": "",
        })

    isi_keywords = ["important safety", "isi", "safety information",
                    "full prescribing", "warnings", "adverse reactions"]
    if any(k in soup.get_text(" ", strip=True).lower() for k in isi_keywords):
        for pct in [25, 50, 75, 100]:
            _add({
                "Page": page_name, "Category": "scroll",
                "Event Name": f"scroll_isi_{pct}", "Action": "scroll",
                "Label": f"ISI {pct}%", "CTA Location": "inpage",
                "CTA Text": f"ISI Scroll {pct}%", "Business Intent": "awareness",
                "Page URL": clean_url, "Event ID": "",
                "Form Info": "", "Video Source": "", "File Type": "",
            })

    # 2. LINKS
    for a in soup.find_all("a", href=True):
        href = a.get("href", "").strip()
        text = a.get_text(strip=True) or href
        if not text or len(text) < 2 or len(text) > 120:
            continue
        full = urljoin(clean_url, href)
        fp   = urlparse(full)
        ext  = os.path.splitext(fp.path)[1].lower()
        loc  = detect_location(a)

        if ext in dl_exts:
            _nh       = a.find_previous(["h2", "h3", "h4"])
            _hslug    = (slug(_nh.get_text(strip=True))[:20] + "_") if _nh else ""
            safe_loc  = _safe_loc(loc, "download")
            ftype     = ext.lstrip(".")
            _add({
                "Page": page_name, "Category": "download",
                "Event Name": f"file_download_{_hslug}{slug(text)[:30]}",
                "Action": "click", "Label": full,
                "CTA Location": safe_loc, "CTA Text": text,
                "Business Intent": "consideration",
                "Page URL": clean_url, "Event ID": "",
                "Form Info": "", "Video Source": "", "File Type": ftype.upper(),
            })
        elif fp.netloc and fp.netloc != parsed.netloc:
            safe_loc = _safe_loc(loc, "exit")
            _add({
                "Page": page_name, "Category": "exit",
                "Event Name": f"exit_click_{slug(text)[:35]}",
                "Action": "click", "Label": full,
                "CTA Location": safe_loc, "CTA Text": text,
                "CTA Type": _infer_cta_type(a),
                "Business Intent": "exit",
                "Page URL": clean_url, "Event ID": "",
                "Form Info": "", "Video Source": "", "File Type": "",
            })
        elif href and not href.startswith("#") and not href.startswith("mailto"):
            loc_suffix = f"_{loc}" if loc in ("header", "footer") else ""
            _add({
                "Page": page_name, "Category": "navigation",
                "Event Name": f"link_click_{slug(text)[:30]}{loc_suffix}",
                "Action": "click", "Label": full,
                "CTA Location": loc, "CTA Text": text,
                "CTA Type": _infer_cta_type(a),
                "Business Intent": "engagement",
                "Page URL": clean_url, "Event ID": "",
                "Form Info": "", "Video Source": "", "File Type": "",
            })

    # 3. FORMS
    for frm in soup.find_all("form"):
        _fid     = frm.get("id", "") or frm.get("name", "") or ""
        _faction = frm.get("action", "") or ""
        _aria    = frm.get("aria-label", "") or ""
        _hs_guid = frm.get("data-form-id","") or frm.get("data-hs-form-guid","")
        _mkto_id = frm.get("id","") if "mkto" in frm.get("id","").lower() else ""
        _nearby_h = ""
        for _tag in ["h1", "h2", "h3", "legend", "h4"]:
            _hel = frm.find(_tag) or frm.find_previous(_tag)
            if _hel:
                _nearby_h = _hel.get_text(strip=True)[:60]
                break
        _submit_btn = (frm.find("button", attrs={"type": "submit"}) or
                       frm.find("input",  attrs={"type": "submit"}) or
                       frm.find("button"))
        _btn_text = ""
        if _submit_btn:
            _btn_text = (_submit_btn.get_text(strip=True) or
                         _submit_btn.get("value", ""))[:40]
        if   _aria:     fname = _aria
        elif _nearby_h: fname = _nearby_h
        elif _hs_guid:  fname = f"hubspot_form_{_hs_guid[:20]}"
        elif _mkto_id:  fname = f"marketo_form_{_mkto_id[:20]}"
        elif _fid and not re.match(r"^[a-f0-9-]{10,}$", _fid.lower()):
            fname = _fid
        elif _faction and "/" in _faction:
            fname = _faction.rstrip("/").split("/")[-1]
        elif _btn_text: fname = _btn_text
        else:           fname = "form"
        fname = str(fname).strip()

        field_labels = []
        for inp in frm.find_all(["input", "select", "textarea"]):
            inp_type = inp.get("type", "text").lower()
            if inp_type in ("hidden", "submit", "button", "reset"):
                continue
            lbl_text = ""
            inp_id = inp.get("id", "")
            if inp_id:
                lbl_el = soup.find("label", attrs={"for": inp_id})
                if lbl_el:
                    lbl_text = lbl_el.get_text(strip=True)
            if not lbl_text:
                lbl_text = (inp.get("placeholder", "") or
                            inp.get("name", "") or
                            inp.get("aria-label", ""))
            if lbl_text:
                field_labels.append(lbl_text[:30])
        form_info = ", ".join(field_labels[:10]) if field_labels else ""

        loc = detect_location(frm)
        for ev, intent in [
            ("form_start",   "acquisition"),
            ("form_submit",  "conversion"),
            ("form_error",   "engagement"),
            ("form_abandon", "engagement"),
        ]:
            _add({
                "Page": page_name, "Category": "form",
                "Event Name": f"{ev}_{_page_slug}",
                "Action": ev, "Label": fname,
                "CTA Location": loc, "CTA Text": _btn_text or fname,
                "Business Intent": intent,
                "Page URL": clean_url, "Event ID": "",
                "Form Info": form_info, "Video Source": "", "File Type": "",
            })

    # iframe-injected forms (HubSpot / Marketo / Eloqua / Pardot)
    for iframe in soup.find_all("iframe"):
        isrc = iframe.get("src", "") or iframe.get("data-src", "")
        if any(k in isrc.lower() for k in ["hubspot","marketo","eloqua","pardot"]):
            _iloc  = detect_location(iframe)
            _inh   = iframe.find_previous(["h2", "h3", "h4"])
            _iname = _inh.get_text(strip=True)[:50] if _inh else "embedded_form"
            for ev, intent in [("form_start","acquisition"),("form_submit","conversion")]:
                _add({
                    "Page": page_name, "Category": "form",
                    "Event Name": f"{ev}_{slug(_iname)[:30]}",
                    "Action": ev, "Label": _iname,
                    "CTA Location": _iloc, "CTA Text": _iname,
                    "Business Intent": intent,
                    "Page URL": clean_url, "Event ID": "",
                    "Form Info": "(embedded iframe form)", "Video Source": "", "File Type": "",
                })

    # 4. VIDEOS
    for v in soup.find_all(["video", "iframe"]):
        vid_source, raw_src = _resolve_video_source(v)
        if v.name == "iframe" and vid_source == "html5":
            continue
        if not (vid_source in ("youtube","vimeo","wistia","brightcove") or v.name == "video"):
            continue

        vname = v.get("title", "").strip()
        if not vname or vname.lower().startswith("http"):
            _vpar = v.find_parent(["section", "article", "div", "figure"])
            if _vpar:
                _vh = _vpar.find(["h1", "h2", "h3", "h4"])
                if _vh:
                    vname = _vh.get_text(strip=True)[:60]
        if not vname:
            vname = page_name

        _vid_loc  = None
        _cta_text = "Watch video"
        for _el in soup.find_all(["button", "a"]):
            _cls = " ".join(_el.get("class", [])).lower()
            _lbl = _el.get("aria-label", "").lower()
            _txt = _el.get_text(strip=True).lower()
            _dat = " ".join([
                _el.get("data-target", ""), _el.get("data-video", ""),
                _el.get("data-src",    ""), _el.get("data-modal", ""),
            ]).lower()
            if any(kw in (_cls + _lbl + _txt + _dat)
                   for kw in ["play","watch","video","vimeo","youtube"]):
                _loc_c = detect_location(_el)
                if _loc_c != "popup":
                    _vid_loc  = _loc_c
                    _cta_text = _el.get_text(strip=True) or "Watch video"
                    break
        if not _vid_loc:
            _vid_loc = detect_location(v) or "inpage"

        for ev in ["video_start", "video_progress_25", "video_progress_50",
                   "video_progress_75", "video_complete", "video_pause"]:
            _add({
                "Page": page_name, "Category": "video",
                "Event Name": ev, "Action": "video",
                "Label": vname, "CTA Location": _vid_loc,
                "CTA Text": _cta_text, "Business Intent": "consideration",
                "Page URL": clean_url, "Event ID": "",
                "Form Info": "", "Video Source": vid_source,
                "File Type": "", "Video Link": raw_src[:120],
            })

    return rows


# -- EXCEL BUILDER --
def build_excel(df, project_name):
    wb   = openpyxl.Workbook()
    hf   = PatternFill("solid", fgColor="1F4E79")
    hfnt = Font(color="FFFFFF", bold=True, size=10)
    af   = PatternFill("solid", fgColor="EBF3FB")
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def make_header(ws, cols, widths):
        ws.row_dimensions[1].height = 26
        for ci, (c, w) in enumerate(zip(cols, widths), 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.fill = hf; cell.font = hfnt
            cell.alignment = ctr; cell.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w

    def add_row(ws, vals, ri):
        fill = af if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=str(v) if v is not None else "")
            cell.fill = fill; cell.font = Font(size=9)
            cell.alignment = lft; cell.border = bdr

    # Sheet 1 — SDR Master
    ws1 = wb.active; ws1.title = "SDR"
    make_header(ws1,
        ["No","Page","Category","Event Name","Action","Label",
         "CTA Location","CTA Text","Business Intent","Page URL","Event ID"],
        [5, 20, 18, 36, 14, 32, 15, 24, 18, 40, 10])
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        add_row(ws1, [
            ri-1,
            row.get("Page",""),       row.get("Category",""),
            row.get("Event Name",""), row.get("Action",""),
            row.get("Label",""),      row.get("CTA Location",""),
            row.get("CTA Text",""),   row.get("Business Intent",""),
            row.get("Page URL",""),   row.get("Event ID",""),
        ], ri)

    # Sheet 2 — Link_Click
    ws2 = wb.create_sheet("Link_Click")
    make_header(ws2,
        ["URL","event","event_name","cta_link","cta_location",
         "cta_text","cta_type","current_page_url","screenshots"],
        [35, 12, 36, 40, 15, 28, 12, 40, 15])
    lc_df = df[df["Category"] == "navigation"]
    for ri, (_, row) in enumerate(lc_df.iterrows(), 2):
        add_row(ws2, [
            row.get("Page URL",""),    row.get("Event ID",""),
            row.get("Event Name",""),  row.get("Label",""),
            row.get("CTA Location",""),row.get("CTA Text",""),
            row.get("CTA Type","link"), row.get("Page URL",""), "",
        ], ri)

    # Sheet 3 — Exit_Click
    ws3 = wb.create_sheet("Exit_Click")
    make_header(ws3,
        ["URL","event","event_name","exit_linkname","exit_url",
         "exit_location","current_page_url","screenshots"],
        [35, 12, 36, 30, 44, 15, 40, 15])
    ec_df = df[df["Category"] == "exit"]
    for ri, (_, row) in enumerate(ec_df.iterrows(), 2):
        add_row(ws3, [
            row.get("Page URL",""),    row.get("Event ID",""),
            row.get("Event Name",""),  row.get("CTA Text",""),
            row.get("Label",""),       row.get("CTA Location",""),
            row.get("Page URL",""),    "",
        ], ri)

    # Sheet 4 — Downloads
    ws4 = wb.create_sheet("Downloads")
    make_header(ws4,
        ["URL","event","event_name","file_text","file_type",
         "file_url","file_location","current_page_url","screenshots"],
        [35, 12, 36, 30, 10, 44, 15, 40, 15])
    dl_df = df[df["Category"] == "download"]
    for ri, (_, row) in enumerate(dl_df.iterrows(), 2):
        href  = row.get("Label","")
        ftype = row.get("File Type", os.path.splitext(href)[1].lstrip(".").upper() or "PDF")
        add_row(ws4, [
            row.get("Page URL",""),    row.get("Event ID",""),
            row.get("Event Name",""),  row.get("CTA Text",""),
            ftype, href,               row.get("CTA Location",""),
            row.get("Page URL",""),    "",
        ], ri)

    # Sheet 5 — Forms
    ws5 = wb.create_sheet("Forms")
    make_header(ws5,
        ["URL","GA4 Event (event)","GA4 Event (event_name)",
         "form_title","form_fields","form_info","current_page_url"],
        [35, 14, 36, 35, 40, 40, 40])
    fm_df = df[df["Category"] == "form"]
    for ri, (_, row) in enumerate(fm_df.iterrows(), 2):
        add_row(ws5, [
            row.get("Page URL",""),    row.get("Event ID",""),
            row.get("Event Name",""),  row.get("Label",""),
            row.get("CTA Text",""),    row.get("Form Info",""),
            row.get("Page URL",""),
        ], ri)

    # Sheet 6 — Videos
    ws6 = wb.create_sheet("Videos")
    make_header(ws6,
        ["URL","GA4_event","Video_name","Video_elapsed_time",
         "Video_link","Video_source","current_page_url"],
        [35, 28, 35, 20, 44, 12, 40])
    vd_df = df[df["Category"] == "video"]
    for ri, (_, row) in enumerate(vd_df.iterrows(), 2):
        ev = row.get("Event Name","")
        if   "progress_25" in ev: elapsed = "25%"
        elif "progress_50" in ev: elapsed = "50%"
        elif "progress_75" in ev: elapsed = "75%"
        elif "complete"    in ev: elapsed = "100%"
        elif "start"       in ev: elapsed = "0%"
        elif "pause"       in ev: elapsed = "variable"
        else:                     elapsed = ""
        add_row(ws6, [
            row.get("Page URL",""),
            row.get("Event Name",""),
            row.get("Label",""),
            elapsed,
            row.get("Video Link", row.get("Label","")),
            row.get("Video Source","html5"),
            row.get("Page URL",""),
        ], ri)

    # Sheet 7 — Scroll
    ws7 = wb.create_sheet("Scroll")
    make_header(ws7,
        ["URL","event","event_name","scroll_depth","current_page_url"],
        [35, 12, 28, 15, 40])
    sc_df = df[df["Category"] == "scroll"]
    for ri, (_, row) in enumerate(sc_df.iterrows(), 2):
        add_row(ws7, [
            row.get("Page URL",""),    row.get("Event ID",""),
            row.get("Event Name",""),  row.get("Label",""),
            row.get("Page URL",""),
        ], ri)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# -- CRAWL HELPER --
def _collect_links_crawl(html_text, seen_set, page_names, netloc_str):
    from urllib.parse import urlparse as _up2
    from bs4 import BeautifulSoup as _BS
    _s = _BS(html_text, "lxml")
    found = []
    skip_exts = {".pdf",".jpg",".jpeg",".png",".gif",".svg",".css",".js",
                 ".zip",".docx",".xlsx",".ppt",".pptx",".mp4",".mp3",".xml"}
    for _a in _s.find_all("a", href=True):
        _href = _a.get("href", "").strip()
        _text = (_a.get("title","") or _a.get_text(strip=True) or "").strip()
        _ph   = _up2(_href)
        if _ph.netloc and _ph.netloc != netloc_str:
            continue
        _path = _ph.path.rstrip("/") or "/"
        if not _path.startswith("/"):
            continue
        if os.path.splitext(_path)[1].lower() in skip_exts:
            continue
        if _path in seen_set:
            continue
        if _text and 1 < len(_text) < 80:
            page_names[_path] = _text[:60]
        elif _path not in page_names:
            page_names[_path] = _path.strip("/").replace("-"," ").replace("_"," ").title() or _path
        found.append(_path)
    return found


# -- STREAMLIT APP --
st.set_page_config(page_title="SDR Generator", layout="wide")

for k, v in {"sdr_df": None, "ga_counter": 4, "project_name": ""}.items():
    if k not in st.session_state:
        st.session_state[k] = v

with st.sidebar:
    st.markdown("### SDR Generator")
    st.caption("3-Step Tracking Plan Builder")
    st.divider()
    if st.session_state.sdr_df is not None and len(st.session_state.sdr_df) > 0:
        df_s = st.session_state.sdr_df
        st.metric("Total Events", len(df_s))
        for cat, grp in df_s.groupby("Category"):
            st.caption(f"{cat}: {len(grp)}")
        st.divider()
    if st.button("Start Over"):
        st.session_state.sdr_df = None
        st.session_state.ga_counter = 4
        st.rerun()

st.title("SDR Generator")
st.caption("Step 1: Define pages  |  Step 2: Review and edit  |  Step 3: Export Excel")
st.divider()

has_draft = st.session_state.sdr_df is not None and len(st.session_state.sdr_df) > 0

with st.expander("Step 1 - Project Setup and Page List", expanded=not has_draft):
    c1, c2 = st.columns(2)
    project_name = c1.text_input("Project / Brand Name",
        value=st.session_state.project_name,
        placeholder="e.g. example")
    site_type = c2.selectbox("Site Type", ["HCP","Patient","Corporate","General"])
    st.session_state.project_name = project_name

    st.markdown("**Choose how to discover pages**")
    disc_tab1, disc_tab2, disc_tab3 = st.tabs([
        "Auto-Crawl", "Sitemap URL", "Paste URL List"])

    with disc_tab1:
        scan_url_input = st.text_input("Root URL",
            placeholder="https://www.example.com or https://user:pass@staging.example.com",
            key="crawl_url")
        st.caption("Crawls the site recursively up to depth 3.")

    with disc_tab2:
        sitemap_url_input = st.text_input("Sitemap URL",
            placeholder="https://www.example.com/sitemap.xml",
            key="sitemap_url")
        st.caption("Parses all URLs from sitemap XML.")

    with disc_tab3:
        manual_url_list = st.text_area("Paste full URLs - one per line", height=160,
            placeholder="https://www.example.com/\nhttps://www.example.com/about/",
            key="manual_urls")
        st.caption("Most accurate - paste the full list from your team.")

    pages_input = st.text_area("Additional pages (Name | /path)", height=80,
        placeholder="Homepage | /\nContact Us | /contact",
        help="Extra pages not covered by discovery above.")

    if st.button("Generate Draft SDR", type="primary", use_container_width=True):
        import requests as _req
        from urllib.parse import urlparse as _up

        all_rows     = []
        ga           = st.session_state.ga_counter
        manual_pages = []

        if pages_input.strip():
            for line in pages_input.strip().splitlines():
                line = line.strip()
                if not line:
                    continue
                if "|" in line:
                    parts = line.split("|", 1)
                    manual_pages.append((parts[0].strip(), parts[1].strip()))
                else:
                    manual_pages.append((line, f"/{slug(line)}"))

        _crawl_input   = st.session_state.get("crawl_url",   "").strip().rstrip("/")
        _sitemap_input = st.session_state.get("sitemap_url", "").strip()
        _manual_input  = st.session_state.get("manual_urls", "").strip()

        base = _crawl_input or _sitemap_input or ""
        if not base and not _manual_input:
            st.warning("Please provide a URL in at least one discovery tab.")
            st.stop()

        if not base and _manual_input:
            _first = _manual_input.strip().splitlines()[0].strip()
            _p2    = _up(_first)
            base   = f"{_p2.scheme}://{_p2.netloc}"

        _parsed   = _up(base)
        _auth_tup = (_parsed.username, _parsed.password) if _parsed.username else None
        _netloc   = _parsed.hostname + (f":{_parsed.port}" if _parsed.port else "")
        _scheme   = _parsed.scheme
        _auth_pfx = f"{_parsed.username}:{_parsed.password}@" if _parsed.username else ""

        def _make_url(path):
            p = path if path.startswith("/") else f"/{path}"
            return f"{_scheme}://{_auth_pfx}{_netloc}{p}"

        _seen        = set()
        scan_targets = []
        _page_names  = {}

        if _manual_input:
            with st.spinner("Loading pages from pasted URL list..."):
                for _u in [u.strip() for u in _manual_input.splitlines() if u.strip()]:
                    try:
                        _pu   = _up(_u)
                        _path = _pu.path or "/"
                        _name = _path.strip("/").replace("-"," ").replace("_"," ").title() or "Homepage"
                        if _u not in [t[1] for t in scan_targets]:
                            scan_targets.append((_name[:60], _u))
                    except Exception:
                        pass
            st.success(f"Loaded {len(scan_targets)} pages from pasted URL list")

        elif _sitemap_input:
            with st.spinner("Parsing sitemap..."):
                try:
                    import xml.etree.ElementTree as _ET
                    _sr = _req.get(_sitemap_input,
                                   headers={"User-Agent":"Mozilla/5.0 Chrome/120"},
                                   timeout=15, auth=_auth_tup)
                    _sr.raise_for_status()
                    _root_xml = _ET.fromstring(_sr.content)
                    _ns       = {"sm":"http://www.sitemaps.org/schemas/sitemap/0.9"}
                    _all_locs = [l.text.strip() for l in _root_xml.findall(".//sm:loc",_ns)]
                    _final_urls = []
                    for _su in _all_locs:
                        if _su.endswith(".xml"):
                            try:
                                _cr = _req.get(_su, headers={"User-Agent":"Mozilla/5.0"},
                                               timeout=10, auth=_auth_tup)
                                _cr_root = _ET.fromstring(_cr.content)
                                for _cl in _cr_root.findall(".//sm:loc",_ns):
                                    _final_urls.append(_cl.text.strip())
                            except Exception:
                                _final_urls.append(_su)
                        else:
                            _final_urls.append(_su)
                    for _u in _final_urls:
                        _pu = _up(_u)
                        if _pu.netloc != _netloc:
                            continue
                        _path = _pu.path or "/"
                        _name = _path.strip("/").replace("-"," ").replace("_"," ").title() or "Homepage"
                        if _u not in [t[1] for t in scan_targets]:
                            scan_targets.append((_name[:60], _u))
                    st.success(f"Found {len(scan_targets)} pages in sitemap")
                except Exception as _se:
                    st.warning(f"Sitemap parse failed: {_se}. Switching to auto-crawl.")

        if not scan_targets and _crawl_input:
            with st.spinner("Auto-discovering pages by crawling..."):
                try:
                    _headers = {"User-Agent":"Mozilla/5.0 Chrome/120"}
                    _resp    = _req.get(_make_url("/"), auth=_auth_tup,
                                        headers=_headers, timeout=15)
                    if _resp.status_code != 200:
                        st.error(f"Could not reach {base} -- HTTP {_resp.status_code}")
                        st.stop()
                    scan_targets.append(("Homepage", _make_url("/")))
                    _seen.add("/")
                    _page_names["/"] = "Homepage"

                    _pass1 = _collect_links_crawl(_resp.text, _seen, _page_names, _netloc)
                    _queue = []
                    for _p in _pass1:
                        if _p not in _seen:
                            _seen.add(_p)
                            _queue.append(_p)

                    _deep = []
                    for _qp in _queue[:60]:
                        try:
                            _r2 = _req.get(_make_url(_qp), auth=_auth_tup,
                                           headers=_headers, timeout=10)
                            if _r2.status_code == 200:
                                for _sp in _collect_links_crawl(_r2.text, _seen, _page_names, _netloc):
                                    if _sp not in _seen:
                                        _seen.add(_sp)
                                        _deep.append(_sp)
                        except Exception:
                            pass

                    _deeper = []
                    for _dp in _deep[:40]:
                        try:
                            _r3 = _req.get(_make_url(_dp), auth=_auth_tup,
                                           headers=_headers, timeout=10)
                            if _r3.status_code == 200:
                                for _sp in _collect_links_crawl(_r3.text, _seen, _page_names, _netloc):
                                    if _sp not in _seen:
                                        _seen.add(_sp)
                                        _deeper.append(_sp)
                        except Exception:
                            pass

                    for _p in _queue + _deep + _deeper:
                        _name = _page_names.get(
                            _p, _p.strip("/").replace("-"," ").replace("_"," ").title() or _p)
                        scan_targets.append((_name[:60], _make_url(_p)))
                    st.success(f"Auto-discovered {len(scan_targets)} pages")
                except Exception as _ce:
                    st.error(f"Crawl failed: {_ce}")
                    st.stop()

        if not scan_targets:
            st.warning("No pages found. Please check your URL or try another discovery method.")
            st.stop()

        for _ep_name, _ep_path in manual_pages:
            _p = _ep_path if _ep_path.startswith("/") else f"/{_ep_path}"
            if _p not in _seen:
                scan_targets.append((_ep_name, _make_url(_p)))
                _seen.add(_p)

        with st.expander(f"{len(scan_targets)} pages discovered - click to preview", expanded=True):
            for _i, (_pn, _pu) in enumerate(scan_targets):
                _clean_disp = _pu.replace(
                    f"{_scheme}://{_auth_pfx}", f"{_scheme}://") if _auth_pfx else _pu
                st.caption(f"{_i+1}. {_pn}  --  {_clean_disp}")

        progress_bar = st.progress(0, text=f"Scanning 0 / {len(scan_targets)} pages...")
        scan_summary = []

        for i, (pname, full_url) in enumerate(scan_targets):
            progress_bar.progress(
                i / len(scan_targets),
                text=f"Scanning {i+1} / {len(scan_targets)}: {pname}...")
            scanned = quick_scan(full_url, page_name=pname)
            count   = 0
            for r in scanned:
                r["Page"]     = pname
                r["Event ID"] = f"ga{ga}"
                ga += 1
                count += 1
            all_rows.extend(scanned)
            scan_summary.append((pname, count))

        progress_bar.progress(1.0,
            text=f"Scan complete -- {len(scan_targets)} pages scanned.")

        total_scanned = sum(c for _, c in scan_summary)
        with st.expander(
                f"Scan Summary -- {total_scanned} elements across {len(scan_targets)} pages",
                expanded=True):
            for pname, count in scan_summary:
                icon = "[OK]" if count > 0 else "[!!]"
                st.caption(f"{icon} {pname}: {count} events")

        if all_rows:
            st.session_state.sdr_df     = pd.DataFrame(all_rows)
            st.session_state.ga_counter = ga
            st.success(
                f"Draft ready -- {len(all_rows)} total events across {len(scan_targets)} page(s)")
            st.rerun()
        else:
            st.warning("No events found. Please check the URL.")

# -- STEP 2: REVIEW & EDIT --
if has_draft:
    with st.expander("Step 2 - Review and Edit Draft SDR", expanded=True):
        st.caption("Edit any cell directly. Save before exporting.")
        COLS = ["Page","Category","Event Name","Action","Label",
                "CTA Location","CTA Text","Business Intent","Page URL","Event ID"]
        df_edit = st.session_state.sdr_df.copy()
        for c in COLS:
            if c not in df_edit.columns:
                df_edit[c] = ""
        edited = st.data_editor(
            df_edit[COLS], use_container_width=True,
            num_rows="dynamic", height=520,
            column_config={
                "Category": st.column_config.SelectboxColumn("Category",
                    options=["navigation","exit","download","form","video",
                             "scroll","conversion","engagement","acquisition"],
                    required=True),
                "Action": st.column_config.SelectboxColumn("Action",
                    options=["click","scroll","submit","play","impression","view","video"],
                    required=False),
                "Business Intent": st.column_config.SelectboxColumn("Business Intent",
                    options=["engagement","conversion","acquisition","consideration",
                             "awareness","support","exit","retention"],
                    required=False),
                "CTA Location": st.column_config.SelectboxColumn("CTA Location",
                    options=["header","footer","inpage","hero","sidebar","popup","banner"],
                    required=False),
            }, key="sdr_editor")

        if st.button("Save Changes", type="secondary"):
            st.session_state.sdr_df = edited.reset_index(drop=True)
            st.success(f"Saved -- {len(edited)} events")
            st.rerun()

    with st.expander("➕ Add Missing Elements (Human-Guided)", expanded=False):
        st.caption("Scanner missed something? Add it manually.")
        mc1, mc2, mc3 = st.columns(3)
        m_type   = mc1.selectbox("Element Type",
                    ["form","button","link","exit","download","video","scroll"])
        m_text   = mc2.text_input("CTA Text / Label", key="m_text")
        m_loc    = mc3.selectbox("CTA Location",
                    ["inpage","header","footer","hero","sidebar","popup"])
        mc4, mc5 = st.columns(2)
        m_intent = mc4.selectbox("Business Intent",
                    ["conversion","acquisition","consideration","engagement",
                     "awareness","support","exit","retention"])
        m_page   = mc5.text_input("Page Name (optional)", key="m_page")
        if st.button("Add to SDR ▶", key="add_manual") and m_text:
            _cat_map = {"form":"form","button":"navigation","link":"navigation",
                        "exit":"exit","download":"download","video":"video","scroll":"scroll"}
            new_row = {
                "Page":           m_page or "Manual",
                "Category":       _cat_map.get(m_type, "navigation"),
                "Event Name":     f"manual_{slug(m_text)[:35]}",
                "Action":         "click",
                "Label":          m_text,
                "CTA Location":   m_loc,
                "CTA Text":       m_text,
                "Business Intent": m_intent,
                "Page URL":       "",
                "Event ID":       f"ga{st.session_state.ga_counter}",
                "Form Info":      "", "Video Source": "", "File Type": "",
            }
            st.session_state.sdr_df = pd.concat(
                [st.session_state.sdr_df, pd.DataFrame([new_row])],
                ignore_index=True)
            st.session_state.ga_counter += 1
            st.success(f"Added: {new_row['Event Name']}")
            st.rerun()

    with st.expander("Step 3 - Export SDR Excel", expanded=True):
        df_exp = st.session_state.sdr_df
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total Events",  len(df_exp))
        c2.metric("Navigation",    len(df_exp[df_exp["Category"]=="navigation"]))
        c3.metric("Forms",         len(df_exp[df_exp["Category"]=="form"]))
        c4.metric("Downloads",     len(df_exp[df_exp["Category"]=="download"]))
        c5.metric("Videos",        len(df_exp[df_exp["Category"]=="video"]))
        c6.metric("Scroll",        len(df_exp[df_exp["Category"]=="scroll"]))

        proj  = st.session_state.project_name or "SDR"
        fname = f"{slug(proj)}_sdr.xlsx"
        buf   = build_excel(df_exp, proj)
        st.download_button(
            label="⬇ Download SDR Excel (7 sheets)",
            data=buf, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
